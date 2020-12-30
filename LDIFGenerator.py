import openpyxl
from colorama import Fore
from colorama import Style
locOutputLDIF="OutputLDIF.ldif"
ldifFile=open(locOutputLDIF,"w")
ldifFile.write("")
ldifFile.close()
ldifFile=open(locOutputLDIF,"a")
def main():
    locInputExcel="LDIFGenerator.xlsx"
    print(f"{Fore.GREEN}Shubham's LDIF Generator{Style.RESET_ALL}")
    workbook=openpyxl.load_workbook(locInputExcel)
    sheet=workbook.active
    max_row=sheet.max_row
    max_column=sheet.max_column
    print("Total Transactions to be processed: {}".format(max_row))
    currentElement=1
    for currentElement in range(1,max_row+1):
        cell=sheet.cell(currentElement,1)
        userID="cn="+cell.value+",ou=users,o=abc"
        cell=sheet.cell(currentElement,2)
        groupName=cell.value
        writeMember(userID,groupName)
        writeGroupMembership(userID,groupName)
        writeEquivalentToMe(userID,groupName)
        if(currentElement%100==0):
            print("{} transactions are processed".format(currentElement))
    workbook.save(locInputExcel)
    ldifFile.close()
    print("{} trancastions are processed".format(currentElement ))

def writeMember(userID,groupName):

    ldifFile.write("dn: {}\nchangetype: modify\nadd: memeber\nmember: {}\n-\n\n".format(groupName,userID))

def writeGroupMembership(userID,groupName):

    ldifFile.write("dn: {}\nchangetype: modify\nadd: groupMembership\ngroupMembership: {}\n-\n\n".format(userID,groupName))

def writeEquivalentToMe(userID,groupName):

    ldifFile.write("dn: {}\nchangetype: modify\nadd: equivalentToMe\nequivalentToMe: {}\n-\n\n".format(groupName,userID))


main()
